import streamlit as st
import pandas as pd
import io
import tempfile
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# --- 模式 1：横向区间模式 ---
def transform_horizontal(df):
    col_to_category = {}
    last_category = "默认分类"
    for col_idx in range(1, df.shape[1]):
        cat_val = df.iloc[1, col_idx] 
        if pd.notna(cat_val) and str(cat_val).strip() not in ["", "分类"]:
            last_category = str(cat_val).strip()
        col_to_category[col_idx] = last_category

    product_names = {}
    for col_idx in range(1, df.shape[1]):
        name_val = df.iloc[2, col_idx]
        if pd.isna(name_val) or str(name_val).strip() == "":
            break 
        product_names[col_idx] = str(name_val).strip()

    results = []
    for i in range(5, len(df)):
        val_a, val_b = df.iloc[i, 0], df.iloc[i, 1]
        if pd.notna(val_a) and pd.notna(val_b):
            purchased_details = []
            row_total_points = 0 
            for col_idx, item_name in product_names.items():
                count = df.iloc[i, col_idx]
                if pd.notna(count) and isinstance(count, (int, float)) and count > 0:
                    category = col_to_category.get(col_idx, "默认分类")
                    row_total_points += int(count)
                    prefix = f"({category})" if category != "默认分类" else ""
                    purchased_details.append(f"{prefix}{item_name}✖{int(count)}")
            
            if purchased_details:
                results.append({
                    "名字": str(val_b).strip(),
                    "（分类名称）/种类✖个数": " / ".join(purchased_details),
                    "总点数": row_total_points,
                    "对应的总金额": str(val_a).strip()
                })
    return pd.DataFrame(results) if results else None

# --- 模式 2：纵向切换模式 ---
def transform_vertical(df):
    item_names = df.iloc[2, 1:].dropna()
    product_map = {col_idx: {"name": str(name).strip()} for col_idx, name in item_names.items()}
    results = []
    current_category = "默认分类"
    for i in range(5, len(df)):
        val_a, val_b = df.iloc[i, 0], df.iloc[i, 1]
        if pd.notna(val_a) and pd.isna(val_b):
            current_category = str(val_a).strip()
            continue
        if pd.notna(val_a) and pd.notna(val_b):
            details = []
            total_pts = 0
            for col_idx, info in product_map.items():
                count = df.iloc[i, col_idx]
                if pd.notna(count) and isinstance(count, (int, float)) and count > 0:
                    total_pts += int(count)
                    prefix = f"({current_category})" if current_category != "默认分类" else ""
                    details.append(f"{prefix}{info['name']}✖{int(count)}")
            if details:
                results.append({
                    "名字": str(val_b).strip(),
                    "（分类名称）/种类×个数": " / ".join(details),
                    "总点数": total_pts,
                    "对应的总金额": str(val_a).strip()
                })
    return pd.DataFrame(results) if results else None

# --- 模式 3：多Sheet合并汇总逻辑 (基于你提供的新代码) ---
def transform_multi_sheet(input_path):
    all_sheets = pd.read_excel(input_path, sheet_name=None)
    result_df = None
    sheet_order = []

    for sheet_name, df in all_sheets.items():
        if df.empty: continue
        sheet_order.append(sheet_name)
        # 统一列名
        df.columns = ['名字', 'list', '点数', '金额'] + list(df.columns[4:])
        
        sheet_summary = df.groupby('名字').agg({
            'list': lambda x: '，'.join(x.astype(str)),
            '点数': 'sum',
            '金额': 'sum'
        }).reset_index()

        sheet_summary.columns = ['名字', f'{sheet_name}_list', f'{sheet_name}_点数', f'{sheet_name}_金额']

        if result_df is None:
            result_df = sheet_summary
        else:
            result_df = result_df.merge(sheet_summary, on='名字', how='outer')

    if result_df is not None:
        amount_cols = [c for c in result_df.columns if c.endswith('_金额')]
        result_df['汇总金额'] = result_df[amount_cols].sum(axis=1, skipna=True)
    
    return result_df, sheet_order

# --- 样式处理 (用于模式3) ---
def apply_excel_style(df, sheet_order):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False)
        
    wb = load_workbook(output)
    ws = wb.active
    ws.insert_rows(1, amount=2)
    ws['A1'] = 'cn'
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    col = 2
    for sheet_name in sheet_order:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        ws.cell(row=1, column=col).value = sheet_name
        ws.cell(row=2, column=col).value = 'list'
        ws.cell(row=2, column=col + 1).value = '点数'
        ws.cell(row=2, column=col + 2).value = '金额'
        col += 4 # 留一列分隔空列

    ws.cell(row=1, column=col).value = '汇总'
    ws.cell(row=2, column=col).value = '金额'
    
    bold_font = Font(bold=True)
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=col).font = bold_font

    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output.getvalue()

# --- Streamlit 界面 ---
st.set_page_config(page_title="制品清单全能工具", layout="wide")
st.title("🛠️ 制品清单全能转换工具")

mode = st.sidebar.radio(
    "请选择转换功能：",
    ("单页转换：横向区间模式", "单页转换：纵向切换模式", "多Sheet合并汇总模式")
)

st.sidebar.info(f"当前模式：{mode}")

uploaded_file = st.file_uploader("上传 Excel 文件 (.xlsx)", type=["xlsx"])

if uploaded_file:
    if mode == "多Sheet合并汇总模式":
        # 需要处理原始文件路径
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name
        
        if st.button("🚀 开始合并"):
            res_df, sheets = transform_multi_sheet(tmp_path)
            if res_df is not None:
                st.success("合并完成！")
                st.dataframe(res_df)
                excel_data = apply_excel_style(res_df, sheets)
                st.download_button("⬇️ 下载合并结果", excel_data, "合并汇总结果.xlsx")
            os.remove(tmp_path)
            
    else:
        # 单页处理模式
        df_raw = pd.read_excel(uploaded_file, header=None)
        if mode == "单页转换：横向区间模式":
            res_df = transform_horizontal(df_raw)
        else:
            res_df = transform_vertical(df_raw)
            
        if res_df is not None:
            st.success("处理成功！")
            st.dataframe(res_df)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                res_df.to_excel(writer, index=False)
            st.download_button("⬇️ 下载转换结果", output.getvalue(), f"转换_{uploaded_file.name}")

